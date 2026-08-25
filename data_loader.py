"""
Loads and validates the 'DB' sheet from an AEB O&M monthly data export
(the same structure as AEB_Data_Asset_for_Pak_Imam.xlsx) into a clean
pandas DataFrame keyed by site.
"""
import pandas as pd
import openpyxl
from column_map import COLUMN_MAP, TEXT_COLS

SENTINEL_CHECKS = [(3, "Company"), (19, "Target"), (55, "Production"), (70, "CO2")]
PCT_KEYS = {key for (_i, key, _l, _g, unit) in COLUMN_MAP if unit == "%"}


class DataLoadError(Exception):
    pass


def _read_header_row2(ws):
    return [c.value for c in ws[2]]


def load_db_workbook(file_like):
    """
    Parameters
    ----------
    file_like : path or file-like object (e.g. Streamlit UploadedFile)

    Returns
    -------
    df : pandas.DataFrame  (one row per site, columns per column_map keys)
    meta : dict with 'period_date' (python date or None) and 'sheet_name'
    """
    try:
        wb = openpyxl.load_workbook(file_like, data_only=True)
    except Exception as exc:
        raise DataLoadError(f"Could not open this file as an Excel workbook: {exc}")

    if "DB" not in wb.sheetnames:
        raise DataLoadError(
            "This workbook has no sheet named 'DB'. Expected the same structure as "
            "the AEB monthly data export (DB / Cycle / Loading % / Coretan)."
        )
    ws = wb["DB"]

    # Sentinel validation against the known AEB DB layout
    row2 = _read_header_row2(ws)
    mismatches = []
    for idx, expected in SENTINEL_CHECKS:
        actual = row2[idx] if idx < len(row2) else None
        if actual != expected:
            mismatches.append(f"column {idx} expected '{expected}', found '{actual}'")
    if mismatches:
        raise DataLoadError(
            "The 'DB' sheet layout does not match the expected AEB template, so this "
            "file cannot be parsed reliably. Differences found: " + "; ".join(mismatches)
        )

    # Reporting period, from cell B1 ("Data per")
    period_date = None
    b1 = ws.cell(row=1, column=2).value
    if hasattr(b1, "year"):
        period_date = b1

    # Data rows: row 4 is the 'Total' row, site rows start at row 5
    records = []
    for row in ws.iter_rows(min_row=5, values_only=True):
        company = row[3] if len(row) > 3 else None
        if company in (None, "", "Total"):
            continue
        rec = {}
        for idx, key, _label, _group, _unit in COLUMN_MAP:
            val = row[idx] if idx < len(row) else None
            rec[key] = val
        records.append(rec)

    if not records:
        raise DataLoadError("No site rows were found below the header — is this an empty or template-only file?")

    df = pd.DataFrame.from_records(records)

    # Coerce numeric columns, leaving text columns as-is
    for idx, key, _label, _group, _unit in COLUMN_MAP:
        if key in TEXT_COLS:
            continue
        df[key] = pd.to_numeric(df[key], errors="coerce")

    # Excel stores percentage-formatted cells as fractions (0-1); normalise to 0-100
    for key in PCT_KEYS:
        if key in df.columns:
            df[key] = df[key] * 100

    df["company"] = df["company"].astype(str).str.strip()
    df["project"] = df["project"].astype(str).str.strip()
    df["estate"] = df["estate"].astype(str).str.strip()
    df["site"] = df["site"].astype(str).str.strip()

    meta = {"period_date": period_date, "n_sites": len(df), "n_companies": df["company"].nunique()}
    return df, meta
